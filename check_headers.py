import openpyxl

try:
    file_path = "C:/Users/sanPedro/Desktop/WorkSpace/followup/audit-automation/Auditorias_Clientes/Casino Talca/11-03/Template Tabla Maestra.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print("--- HEADERS ROW 1 ---")
    for i in range(1, 26):
        val = ws.cell(row=1, column=i).value
        print(f"Col {i}: {val}")
        
    print("\n--- HEADERS ROW 2 ---")
    headers = {}
    for i in range(1, 26):
        val = ws.cell(row=2, column=i).value
        print(f"Col {i}: {val}")
        if val:
            headers[str(val).strip()] = i
            
    print("\n--- DICT MAPPING ---")
    print(headers)
except Exception as e:
    print(f"ERROR: {e}")
