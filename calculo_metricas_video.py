import pandas as pd
import argparse
import os
import io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def process_audit_data(empresa, fecha, sucursal=None, input_filename='input.csv'):
    # Construir rutas dinámicamente
    base_dir = "Auditorias_Clientes"
    
    if sucursal:
        work_dir = os.path.join(base_dir, empresa, sucursal, fecha)
    else:
        work_dir = os.path.join(base_dir, empresa, fecha)
        
    input_file_path = os.path.join(work_dir, input_filename)

    # Validar que el archivo y directorio existan
    if not os.path.exists(input_file_path):
        return None, f"❌ ERROR: No se encontró el archivo en la ruta:\n   {input_file_path}"

    # --- PASO 1: Leer el archivo ---
    try:
        df = pd.read_csv(input_file_path)
    except Exception as e:
        return None, f"❌ ERROR al leer CSV: {str(e)}"

    # Filtrar solo registros que tengan un 'Zona_name' asignado
    df_zones = df[df['Zona_name'].notna() & (df['Zona_name'] != '')].copy()
    
    # PASO 2.0: Configuración Clave Única
    has_camera = 'Camara' in df_zones.columns
    group_cols = ['Camara', 'Zona_name'] if has_camera else ['Zona_name']
    
    def calculate_metrics(group):
        total_eventos = len(group)
        eventos_no_registrados = group['Identity_ID'].isna().sum() + \
                                (group['Identity_ID'].astype(str).str.strip() == '').sum()
        eventos_registrados = total_eventos - eventos_no_registrados
        precicion_eventos = group['Event_Audit'].astype(str).str.strip().str.lower().eq('bien').sum()
        precision_genero = group['Gender_Audit'].astype(str).str.strip().str.lower().eq('bien').sum()
        precision_edad = group['Age_Audit'].astype(str).str.strip().str.lower().eq('bien').sum()
        genero_conocido = group['Gender'].notna() & (group['Gender'] != 'unknown')
        cobertura_genero = genero_conocido.sum()
        edad_numerica = pd.to_numeric(group['Age'], errors='coerce')
        edad_conocida = edad_numerica.notna() & (edad_numerica > 0)
        cobertura_edad = edad_conocida.sum()
        identity_unknown = (group['Identity_ID'] == 'unknown').sum()
        cobertura_identity = eventos_registrados - identity_unknown
        
        # Nueva métrica: Registrados pero Incorrectos
        eventos_correctos = precicion_eventos
        eventos_registrados_mal = eventos_registrados - eventos_correctos

        def calc_pct(part, total):
            return (part / total) if total > 0 else 0.0
            
        return pd.Series({
            'Total Eventos': total_eventos,
            'Eventos Registrados por el Sistema': eventos_registrados,
            '% Eventos Registrados por el Sistema': calc_pct(eventos_registrados, total_eventos),
            'Eventos NO Registrados (Manuales)': eventos_no_registrados,
            '% Eventos NO Registrados (Manuales)': calc_pct(eventos_no_registrados, total_eventos),
            'Eventos Correctos del Sistema': eventos_correctos,
            '% Eventos Correctos del Sistema': calc_pct(eventos_correctos, total_eventos),
            'Eventos Reg. Mal (Sist.)': eventos_registrados_mal,
            '% Eventos Reg. Mal (Sist.)': calc_pct(eventos_registrados_mal, eventos_registrados),
            'Cobertura Identity': cobertura_identity,
            '% Cobertura Identity': calc_pct(cobertura_identity, eventos_registrados),
            'Identity Unknown': identity_unknown,
            '% Identity Unknown': calc_pct(identity_unknown, eventos_registrados),
            'Cobertura Género': cobertura_genero,
            '% Cobertura Género': calc_pct(cobertura_genero, eventos_registrados),
            'Precisión de Género': precision_genero,
            '% Precisión de Género': calc_pct(precision_genero, cobertura_genero) if cobertura_genero > 0 else "N/A",
            'Cobertura Edad': cobertura_edad,
            '% Cobertura Edad': calc_pct(cobertura_edad, eventos_registrados),
            'Precisión de Edad': precision_edad,
            '% Precisión de Edad': calc_pct(precision_edad, cobertura_edad) if cobertura_edad > 0 else "N/A"
        })

    # Agrupar por zona
    reporte = df_zones.groupby(group_cols).apply(calculate_metrics, include_groups=False).reset_index()

    if has_camera:
        reporte.rename(columns={'Zona_name': 'Zona', 'Camara': 'Camara_pura'}, inplace=True)
    else:
        reporte.rename(columns={'Zona_name': 'Zona'}, inplace=True)
        reporte['Camara_pura'] = ""

    reporte['Fecha'] = fecha.replace('_', '-')
    reporte['Hora_inicio'] = ""; reporte['Hora_termino'] = ""

    # Totales
    totales_res = calculate_metrics(df_zones)
    totales_res['Zona'] = 'TOTAL'; totales_res['Camara_pura'] = ""
    totales_res['Fecha'] = fecha.replace('_', '-')
    totales_res['Hora_inicio'] = ""; totales_res['Hora_termino'] = ""
    reporte = pd.concat([reporte, pd.DataFrame(totales_res).T], ignore_index=True)

    # Reordenar Columnas para Consistencia Dashboard/Excel (Sincronizado con Plantilla)
    cols_order = [
        'Zona', 'Fecha', 'Hora_inicio', 'Hora_termino', 'Camara_pura',
        'Total Eventos', 'Eventos Registrados por el Sistema', '% Eventos Registrados por el Sistema',
        'Eventos NO Registrados (Manuales)', '% Eventos NO Registrados (Manuales)',
        'Eventos Correctos del Sistema', '% Eventos Correctos del Sistema',
        'Eventos Reg. Mal (Sist.)', '% Eventos Reg. Mal (Sist.)',
        'Cobertura Identity', '% Cobertura Identity', 'Identity Unknown', '% Identity Unknown',
        'Cobertura Género', '% Cobertura Género', 'Precisión de Género', '% Precisión de Género',
        'Cobertura Edad', '% Cobertura Edad', 'Precisión de Edad', '% Precisión de Edad'
    ]
    # Filtrar solo las que existan para evitar errores si alguna falta
    cols_order = [c for c in cols_order if c in reporte.columns]
    reporte = reporte[cols_order]

    # --- GENERACIÓN DE GRÁFICOS (Siempre se generan para la UI) ---
    df_grafico = reporte[reporte['Zona'] != 'TOTAL'].copy()
    df_total   = reporte[reporte['Zona'] == 'TOTAL'].copy()
    
    def to_num(series): 
        # Convertir a numérico, manejando strings que traen '%' si fuera necesario
        s = series.astype(str).str.replace('%', '', regex=False)
        return pd.to_numeric(s, errors='coerce').fillna(0)

    # --- FUNCIÓN HELPER: Gráfico de Doble Barra (Estilo Crystal Pro) ---
    def _make_dual_bar_chart(zonas, base_vals, result_vals, titulo, label_base="Auditados", label_result="Correctos"):
        n = len(zonas)
        _BAR_W = 0.30
        _BAR_GAP = 0.05
        _HALF = _BAR_W / 2 + _BAR_GAP / 2
        
        fig_w = max(6, min(16, 3.0 + n * 1.5))
        fig, ax = plt.subplots(figsize=(fig_w, 6))
        fig.patch.set_facecolor('white')
        
        # Lógica Crystal: La barra 1 es la BASE (100% de referencia)
        # La barra 2 es el ÉXITO (% relativo a la base)
        heights1 = [100.0 for _ in range(n)]
        heights2 = [(r / b * 100) if b > 0 else 0 for r, b in zip(result_vals, base_vals)]
        
        x = range(n)
        # Barra de base (Azul oscuro - Referencia 100%)
        rects1 = ax.bar([i - _HALF for i in x], heights1, _BAR_W, label=label_base, color='#1B2A4A', zorder=3)
        # Barra de resultado (Rojo - % de cumplimiento)
        rects2 = ax.bar([i + _HALF for i in x], heights2, _BAR_W, label=label_result, color='#C0392B', zorder=3)
        
        # Cuadrícula horizontal muy clara (Líneas hacia la derecha)
        ax.grid(axis='y', linestyle='-', color='#EEEEEE', alpha=0.8, zorder=0)
        
        # Etiquetas: MOSTRAR NÚMEROS EXACTOS SOBRE LAS BARRAS
        def autolabel(rects, vals):
            for rect, val in zip(rects, vals):
                h = rect.get_height()
                ax.annotate(f'{int(val)}',
                            xy=(rect.get_x() + rect.get_width() / 2, h),
                            xytext=(0, 5),
                            textcoords="offset points",
                            ha='center', va='bottom', fontsize=10, fontweight='bold', color='#1B2A4A')
        
        autolabel(rects1, base_vals)
        autolabel(rects2, result_vals)
        
        ax.set_ylim(0, 115) # Espacio para las etiquetas superiores
        ax.yaxis.set_major_formatter(mticker.PercentFormatter())
        ax.set_xticks(x)
        ax.set_xticklabels(zonas, rotation=30, ha='right', fontsize=9, fontweight='medium')
        ax.set_xlim(-0.7, n - 0.3)
        
        # Estética Crystal: Bordes limpios
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color('#BBBBBB')
        ax.spines['bottom'].set_color('#BBBBBB')
        
        ax.set_title(titulo, pad=60, fontweight='bold', color='#1B2A4A', fontsize=16)
        ax.legend(loc='upper center', bbox_to_anchor=(0.5, 1.15), ncol=2, frameon=False, fontsize=11)
        
        plt.tight_layout()
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=110, bbox_inches='tight')
        plt.close(fig)
        return buf.getvalue()

    # 1. Gráfico Global
    zonas_labels = df_grafico['Zona'].astype(str).tolist()
    t_vals = to_num(df_grafico['Total Eventos']).tolist()
    p_vals = to_num(df_grafico['Eventos Correctos del Sistema']).tolist()
    
    img_global_bytes = _make_dual_bar_chart(zonas_labels, t_vals, p_vals, "Eventos Correctos por Zona", "Total Auditados", "Correctos")
    img_global = io.BytesIO(img_global_bytes)

    # 2. Resumen Totales
    img_totales = None
    img_totales_bytes = None
    if not df_total.empty:
        t_row = df_total.iloc[0]
        fig2, ax2 = plt.subplots(figsize=(6, 5))
        cats = ['Total', 'Registrados', 'Correctos']
        vals = [to_num(pd.Series([v]))[0] for v in [t_row['Total Eventos'], t_row['Eventos Registrados por el Sistema'], t_row['Eventos Correctos del Sistema']]]
        ax2.bar(cats, vals, color=['#1B2A4A', '#2C5282', '#C0392B'], width=0.5)
        ax2.set_title("Resumen Global", pad=20, fontweight='bold')
        plt.tight_layout()
        _buf2 = io.BytesIO(); plt.savefig(_buf2, format='png', dpi=120); plt.close(fig2)
        img_totales_bytes = _buf2.getvalue()
        img_totales = io.BytesIO(img_totales_bytes)

    # 3. Por Cámara (Resumen Agregado, Precisión y Cobertura)
    cam_images_list = []
    cam_coverage_images_list = []
    cam_summary_images_list = []
    cameras = df_grafico['Camara_pura'].unique()
    for cam in cameras:
        cam_df = df_grafico[df_grafico['Camara_pura'] == cam]
        if cam_df.empty: continue
        cam_label = f"Cámara {int(cam)}" if cam not in ('', 'nan', None) else "General"
        z_cam = cam_df['Zona'].tolist()
        t_c = to_num(cam_df['Total Eventos']).tolist()
        r_c = to_num(cam_df['Eventos Registrados por el Sistema']).tolist()
        p_c = to_num(cam_df['Eventos Correctos del Sistema']).tolist()
        
        # A. Gráfico de Resumen Agregado (Petición Audio: Total Cámara vs Precisión Cámara)
        t_cam_sum = sum(t_c)
        p_cam_sum = sum(p_c)
        cam_sum_bytes = _make_dual_bar_chart(["TOTAL CÁMARA"], [t_cam_sum], [p_cam_sum], f"Resumen Agregado — {cam_label}", "Total Auditado", "Correctos")
        cam_summary_images_list.append({'label': cam_label, 'buffer': io.BytesIO(cam_sum_bytes), 'bytes': cam_sum_bytes})

        # B. Gráfico de Precisión (Auditados vs Correctos por Zona)
        cam_bytes = _make_dual_bar_chart(z_cam, t_c, p_c, f"Eventos Correctos por Zona — {cam_label}", "Global Auditados", "Correctos")
        cam_images_list.append({'label': cam_label, 'buffer': io.BytesIO(cam_bytes), 'bytes': cam_bytes})
        
        # C. Gráfico de Cobertura (Total Auditado vs Registrados por Sistema por Zona)
        cov_bytes = _make_dual_bar_chart(z_cam, t_c, r_c, f"Cobertura por Zona — {cam_label}", "Total Reales", "Sist. Captados")
        cam_coverage_images_list.append({'label': cam_label, 'buffer': io.BytesIO(cov_bytes), 'bytes': cov_bytes})

    # 4. Detalle por Zona (Individual)
    zone_images_list = []
    for i, row in df_grafico.iterrows():
        z_name = str(row['Zona'])
        t_z = [to_num(pd.Series([row['Total Eventos']]))[0]]
        p_z = [to_num(pd.Series([row['Eventos Correctos del Sistema']]))[0]]
        
        z_bytes = _make_dual_bar_chart([z_name], t_z, p_z, f"Desglose Correctos: {z_name}", "Base Auditada", "Correctos")
        zone_images_list.append({'label': z_name, 'buffer': io.BytesIO(z_bytes), 'bytes': z_bytes})

    # 5. Análisis de Unknowns (Identidades Desconocidas)
    # A. Gráfico Consolidado (Global)
    t_reg_global = df_total.iloc[0]['Eventos Registrados por el Sistema']
    t_unk_global = df_total.iloc[0]['Identity Unknown']
    img_unknown_global_bytes = _make_dual_bar_chart(["TOTAL SITIO"], [t_reg_global], [t_unk_global], "Impacto Global de Unknowns", "Sist. Registrados", "Ident. Unknown")
    img_unknown_global = io.BytesIO(img_unknown_global_bytes)

    # B. Gráficos por Zona
    unknown_images_list = []
    for i, row in df_grafico.iterrows():
        z_name = str(row['Zona'])
        reg_z = [to_num(pd.Series([row['Eventos Registrados por el Sistema']]))[0]]
        unkn_z = [to_num(pd.Series([row['Identity Unknown']]))[0]]
        
        if reg_z[0] > 0:
            unkn_bytes = _make_dual_bar_chart([z_name], reg_z, unkn_z, f"Impacto Unknowns: {z_name}", "Sist. Registrados", "Unknowns")
            unknown_images_list.append({'label': z_name, 'buffer': io.BytesIO(unkn_bytes), 'bytes': unkn_bytes})

    # --- EXPORTACIÓN EXCEL ---
    # Priorizar la plantilla centralizada para mantener consistencia de formato
    central_template = os.path.join(os.getcwd(), 'templates', 'Template Tabla Maestra.xlsx')
    local_template = os.path.join(work_dir, 'Template Tabla Maestra.xlsx')
    template_path = central_template if os.path.exists(central_template) else local_template
    output_filename = os.path.join(work_dir, f'Reporte_Auditoria_Maestro_{empresa}.xlsx')
    
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path); ws = wb.active
        col_mapping = {
            'Zona': 1, 'Fecha': 2, 'Hora_inicio': 3, 'Hora_termino': 4, 'Camara_pura': 5,
            'Total Eventos': 6, 'Eventos Registrados por el Sistema': 7, '% Eventos Registrados por el Sistema': 8,
            'Eventos NO Registrados (Manuales)': 9, '% Eventos NO Registrados (Manuales)': 10,
            'Eventos Correctos del Sistema': 11, '% Eventos Correctos del Sistema': 12,
            'Eventos Reg. Mal (Sist.)': 13, '% Eventos Reg. Mal (Sist.)': 14,
            'Cobertura Identity': 15, '% Cobertura Identity': 16,
            'Identity Unknown': 17, '% Identity Unknown': 18,
            'Cobertura Género': 19, '% Cobertura Género': 20,
            'Precisión de Género': 21, '% Precisión de Género': 22,
            'Cobertura Edad': 23, '% Cobertura Edad': 24,
            'Precisión de Edad': 25, '% Precisión de Edad': 26
        }
        # --- ESTILOS DE TABLA MAESTRA ---
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        grey_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        # 0. Limpiar datos previos (desde fila 3 hasta 500 para asegurar)
        for row_to_clear in range(3, 500):
            for col_to_clear in range(1, 25):
                ws.cell(row=row_to_clear, column=col_to_clear).value = None
                ws.cell(row=row_to_clear, column=col_to_clear).fill = PatternFill(fill_type=None)
                ws.cell(row=row_to_clear, column=col_to_clear).border = Border()

        # 1. Cabeceras (Fila 2): Letras Blancas y Negrita
        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        for col_idx in range(1, 27):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # 2. Filas de Datos y TOTAL
        for i, row in reporte.iterrows():
            curr_row = 3 + i
            is_total_row = str(row['Zona']).upper() == 'TOTAL'
            
            # Definir fuente: TOTAL es más grande (11) y Negrita
            if is_total_row:
                row_font = Font(name='Arial', size=11, bold=True)
            else:
                row_font = Font(name='Arial', size=10)
            
            for df_col, ex_col in col_mapping.items():
                if df_col in reporte.columns:
                    val = row[df_col]
                    cell = ws.cell(row=curr_row, column=ex_col)
                    
                    # Columnas de Porcentaje -> Fórmulas dinámicas
                    if isinstance(df_col, str) and df_col.startswith('%'):
                        formulas = {
                            '% Eventos Registrados por el Sistema': f'=IF(F{curr_row}>0, G{curr_row}/F{curr_row}, 0)',
                            '% Eventos NO Registrados (Manuales)': f'=IF(F{curr_row}>0, I{curr_row}/F{curr_row}, 0)',
                            '% Eventos Correctos del Sistema': f'=IF(F{curr_row}>0, K{curr_row}/F{curr_row}, 0)',
                            '% Eventos Reg. Mal (Sist.)': f'=IF(G{curr_row}>0, M{curr_row}/G{curr_row}, 0)',
                            '% Cobertura Identity': f'=IF(G{curr_row}>0, O{curr_row}/G{curr_row}, 0)',
                            '% Identity Unknown': f'=IF(G{curr_row}>0, Q{curr_row}/G{curr_row}, 0)',
                            '% Cobertura Género': f'=IF(G{curr_row}>0, S{curr_row}/G{curr_row}, 0)',
                            '% Precisión de Género': f'=IF(S{curr_row}>0, U{curr_row}/S{curr_row}, "")',
                            '% Cobertura Edad': f'=IF(G{curr_row}>0, W{curr_row}/G{curr_row}, 0)',
                            '% Precisión de Edad': f'=IF(W{curr_row}>0, Y{curr_row}/W{curr_row}, "")'
                        }
                        if df_col in formulas:
                            cell.value = formulas[df_col]
                        else:
                            try: cell.value = float(val) if val not in (None, "", "N/A") else 0
                            except: cell.value = 0
                        cell.number_format = '0.00%'
                    
                    # Columnas de Conteo (Enteros)
                    elif df_col in ['Total Eventos', 'Eventos Registrados por el Sistema', 'Eventos NO Registrados (Manuales)', 'Eventos Correctos del Sistema', 'Precisión de Género', 'Precisión de Edad', 'Eventos Reg. Mal (Sist.)', 'Cobertura Género', 'Cobertura Edad', 'Cobertura Identity', 'Identity Unknown']:
                        try:
                            cell.value = int(float(val)) if val not in (None, "", "N/A") else 0
                        except:
                            cell.value = 0
                        cell.number_format = '0'
                    
                    # Texto / Otros
                    else:
                        cell.value = val if val is not None else ""
                        cell.number_format = 'General'
                        
                    # Estilos comunes
                    cell.font = row_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    cell.fill = grey_fill
        ws_g = wb.create_sheet("Gráficos")
        ws_g.add_image(XLImage(io.BytesIO(img_global_bytes)), "B2")
        if img_totales_bytes: ws_g.add_image(XLImage(io.BytesIO(img_totales_bytes)), "B30")
        ws_z = wb.create_sheet("Por Cámara")
        cursor = 2
        for i, img_obj in enumerate(cam_images_list):
            # 1. Resumen Agregado (Arriba)
            if i < len(cam_summary_images_list):
                ws_z.add_image(XLImage(io.BytesIO(cam_summary_images_list[i]['bytes'])), f"B{cursor}")
                cursor += 25
            
            # 2. Precisión y Cobertura (Lado a lado)
            ws_z.add_image(XLImage(io.BytesIO(img_obj['bytes'])), f"B{cursor}")
            if i < len(cam_coverage_images_list):
                ws_z.add_image(XLImage(io.BytesIO(cam_coverage_images_list[i]['bytes'])), f"L{cursor}")
            cursor += 28
            
        ws_det = wb.create_sheet("Detalle por Zona")
        cursor = 2
        for img_obj in zone_images_list:
            ws_det.add_image(XLImage(io.BytesIO(img_obj['bytes'])), f"B{cursor}")
            cursor += 25

        ws_unk = wb.create_sheet("Análisis de Unknowns")
        ws_unk.add_image(XLImage(io.BytesIO(img_unknown_global_bytes)), "B2")
        cursor = 28
        for img_obj in unknown_images_list:
            ws_unk.add_image(XLImage(io.BytesIO(img_obj['bytes'])), f"B{cursor}")
            cursor += 25
        wb.save(output_filename)
    else:
        output_filename = os.path.join(work_dir, f'Reporte_Auditoria_{empresa}_{fecha}.xlsx')
        reporte.to_excel(output_filename, index=False)
        
    return {
        'reporte': reporte,
        'df_grafico': df_grafico,
        'df_total': df_total,
        'img_global': img_global,
        'img_totales': img_totales,
        'cam_images': cam_images_list,
        'cam_coverage_images': cam_coverage_images_list,
        'cam_summary_images': cam_summary_images_list,
        'zone_images': zone_images_list,
        'img_unknown_global': img_unknown_global,
        'img_unknown_global_bytes': img_unknown_global_bytes,
        'unknown_images': unknown_images_list,
        'output_xlsx': output_filename,
        'work_dir': work_dir
    }, None

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('-e', '--empresa', required=True)
    parser.add_argument('-f', '--fecha', required=True)
    parser.add_argument('-i', '--input', default='input.csv')
    args = parser.parse_args()
    res, err = process_audit_data(args.empresa, args.fecha, args.input)
    if err: print(err); exit(1)
    print(f"✅ Reporte generado: {res['output_xlsx']}")
