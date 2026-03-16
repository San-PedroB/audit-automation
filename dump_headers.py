import openpyxl

file_path = "c:/Users/sanPedro/Desktop/WorkSpace/followup/audit-automation/Auditorias_Clientes/Casino Talca/11-03/Template_Tabla_Maestra.xlsx"
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    headers_row_1 = []
    headers_row_2 = []
    
    for col in range(1, 30):
        val1 = ws.cell(row=1, column=col).value
        val2 = ws.cell(row=2, column=col).value
        
        headers_row_1.append(str(val1) if val1 is not None else "")
        headers_row_2.append(str(val2) if val2 is not None else "")
        
    print("ROW 1:", headers_row_1)
    print("ROW 2:", headers_row_2)
except Exception as e:
    print(f"Error: {e}")
