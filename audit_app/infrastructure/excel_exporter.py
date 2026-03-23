"""Excel export infrastructure."""

from __future__ import annotations

import io
import os

import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from audit_app.domain.kpi_schema import COUNT_COLUMNS, KPI_COLUMNS


def _write_df_to_ws(ws, df: pd.DataFrame, start_row: int, start_col: int):
    if df is None or df.empty:
        return
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Fill empty columns or convert arrays to strings if needed
    clean_df = df.copy()
    for col in clean_df.columns:
        if str(col).startswith("%"):
            clean_df[col] = pd.to_numeric(clean_df[col], errors="coerce").fillna(0.0)
            
    rows = list(dataframe_to_rows(clean_df, index=False, header=True))
    for r_idx, row_data in enumerate(rows):
        for c_idx, value in enumerate(row_data):
            cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)
            cell.border = thin_border
            if r_idx == 0:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Format percentages
            if r_idx > 0 and isinstance(clean_df.columns[c_idx], str) and clean_df.columns[c_idx].startswith("%"):
                cell.number_format = "0.00%"


def _format_summary_df(df_source: pd.DataFrame) -> pd.DataFrame:
    if df_source is None or df_source.empty:
        return pd.DataFrame()
        
    summary_rows = []
    for _, row in df_source.iterrows():
        total_eventos = float(row.get("Total Eventos", 0)) if pd.notna(row.get("Total Eventos")) else 0
        reg_sistema = float(row.get("Eventos Registrados por el Sistema", 0)) if pd.notna(row.get("Eventos Registrados por el Sistema")) else 0
        correct_sistema = float(row.get("Eventos Correctos del Sistema", 0)) if pd.notna(row.get("Eventos Correctos del Sistema")) else 0
        no_reg = float(row.get("Eventos NO Registrados (Manuales)", 0)) if pd.notna(row.get("Eventos NO Registrados (Manuales)")) else 0
        
        pct_correct = row.get("% Eventos Correctos sobre Registrados", 0)
        pct_correct = float(pct_correct) if pd.notna(pct_correct) else 0.0

        try:
            x_val = int(round(correct_sistema))
            y_val = int(round(reg_sistema))
            precision_str = f"{x_val} de {y_val} ({pct_correct:.2%})"
        except Exception:
            precision_str = "0 de 0 (0.00%)"
            
        summary_rows.append({
            "Zona": row.get("Zona", "Desconocida"),
            "Total eventos": int(round(total_eventos)),
            "Eventos registrados por el sistema": int(round(reg_sistema)),
            "Precisión sobre registrados": precision_str,
            "Eventos no registrados por el sistema": int(round(no_reg))
        })
    return pd.DataFrame(summary_rows)


def export_audit_report(
    empresa: str,
    fecha: str,
    work_dir: str,
    reporte: pd.DataFrame,
    chart_payloads: dict,
    df_grafico: pd.DataFrame | None = None,
    df_total: pd.DataFrame | None = None,
    output_filename: str | None = None,
) -> str:
    central_template = os.path.join(os.getcwd(), "templates", "Template Tabla Maestra.xlsx")
    local_template = os.path.join(work_dir, "Template Tabla Maestra.xlsx")
    template_path = central_template if os.path.exists(central_template) else local_template
    if not output_filename:
        output_filename = os.path.join(work_dir, f"Reporte_Auditoria_Maestro_{empresa}.xlsx")

    if not os.path.exists(template_path):
        fallback_filename = os.path.join(work_dir, f"Reporte_Auditoria_{empresa}_{fecha}.xlsx")
        reporte.to_excel(fallback_filename, index=False)
        return fallback_filename

    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active
    column_mapping = {column_name: index + 1 for index, column_name in enumerate(KPI_COLUMNS)}

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    for column_index in range(1, len(KPI_COLUMNS) + 1):
        cell = worksheet.cell(row=2, column=column_index)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row_index, row in reporte.iterrows():
        excel_row = 3 + row_index
        is_total_row = str(row["Zona"]).upper() == "TOTAL"
        row_font = Font(name="Arial", size=11, bold=True) if is_total_row else Font(name="Arial", size=10)

        for column_name, column_index in column_mapping.items():
            if column_name not in reporte.columns:
                continue

            value = row[column_name]
            cell = worksheet.cell(row=excel_row, column=column_index)

            if isinstance(column_name, str) and column_name.startswith("%"):
                try:
                    cell.value = float(value) if value not in (None, "", "N/A") else 0
                except Exception:
                    cell.value = 0
                cell.number_format = "0.00%"
            elif column_name in COUNT_COLUMNS:
                try:
                    cell.value = int(float(value)) if value not in (None, "", "N/A") else 0
                except Exception:
                    cell.value = 0
                cell.number_format = "0"
            elif column_name == "Tipo_sensor":
                cell.value = str(value) if value is not None else ""
                cell.number_format = "General"
            else:
                cell.value = value if value is not None else ""
                cell.number_format = "General"

            cell.font = row_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.fill = grey_fill

    from audit_app.domain.kpi_schema import CAMERA_COLUMN

    ws_graphs = workbook.create_sheet("Graficos")
    ws_graphs.add_image(XLImage(io.BytesIO(chart_payloads["img_global_bytes"])), "B2")
    if df_grafico is not None:
        _write_df_to_ws(ws_graphs, df_grafico, 2, 12)
        
    if chart_payloads["img_totales_bytes"]:
        ws_graphs.add_image(XLImage(io.BytesIO(chart_payloads["img_totales_bytes"])), "B30")
        if df_total is not None:
            _write_df_to_ws(ws_graphs, df_total, 30, 12)

    ws_camera = workbook.create_sheet("Por Camara")
    cursor = 2
    for index, img_obj in enumerate(chart_payloads["cam_images"]):
        cam_label = img_obj["label"]
        if index < len(chart_payloads["cam_summary_images"]):
            ws_camera.add_image(
                XLImage(io.BytesIO(chart_payloads["cam_summary_images"][index]["bytes"])),
                f"B{cursor}",
            )
            cursor += 25

        ws_camera.add_image(XLImage(io.BytesIO(img_obj["bytes"])), f"B{cursor}")
        
        if df_grafico is not None:
            # We match by looking up "General" mappings back to missing values or exact matches.
            # Simplified matching using the raw df_grafico filtered by format logic
            def format_cam(val):
                if val in ("", "nan", None): return "General"
                try: return f"Camara {int(val)}"
                except: return f"Camara {val}"
            
            cam_df = df_grafico[df_grafico[CAMERA_COLUMN].apply(format_cam) == cam_label]
            _write_df_to_ws(ws_camera, _format_summary_df(cam_df), cursor, 24)

        if index < len(chart_payloads["cam_coverage_images"]):
            ws_camera.add_image(
                XLImage(io.BytesIO(chart_payloads["cam_coverage_images"][index]["bytes"])),
                f"L{cursor}",
            )
        cursor += 28

    ws_zone = workbook.create_sheet("Detalle por Zona")
    cursor = 2
    for img_obj in chart_payloads["zone_images"]:
        zone_name = img_obj["label"]
        ws_zone.add_image(XLImage(io.BytesIO(img_obj["bytes"])), f"B{cursor}")
        
        if df_grafico is not None:
            zone_df = df_grafico[df_grafico["Zona"].astype(str) == zone_name].copy()
            _write_df_to_ws(ws_zone, _format_summary_df(zone_df), cursor, 12)

        cursor += 25

    ws_unknown = workbook.create_sheet("Analisis de Unknowns")
    ws_unknown.add_image(XLImage(io.BytesIO(chart_payloads["img_unknown_global_bytes"])), "B2")
    
    if df_total is not None:
        unknown_cols = ["Zona", "Eventos Registrados por el Sistema", "Identity Unknown", "% Identity Unknown"]
        avail_u_cols = [c for c in unknown_cols if c in df_total.columns]
        if avail_u_cols:
            _write_df_to_ws(ws_unknown, df_total[avail_u_cols], 2, 12)

    cursor = 28
    for img_obj in chart_payloads["unknown_images"]:
        zone_name = img_obj["label"]
        ws_unknown.add_image(XLImage(io.BytesIO(img_obj["bytes"])), f"B{cursor}")
        
        if df_grafico is not None:
            zone_u_df = df_grafico[df_grafico["Zona"].astype(str) == zone_name].copy()
            avail_u_cols = [c for c in unknown_cols if c in zone_u_df.columns]
            if avail_u_cols:
                _write_df_to_ws(ws_unknown, zone_u_df[avail_u_cols], cursor, 12)
        
        cursor += 25

    workbook.save(output_filename)
    return output_filename
